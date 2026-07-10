from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import os
import io
import pandas as pd
import google.generativeai as genai
from fastapi.responses import StreamingResponse
import json
import re

app = FastAPI(title="CreativeOps AI Engine", version="2.1")

# Cấu hình CORS để Front-end và Back-end giao tiếp không bị chặn
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- CẤU HÌNH GOOGLE GEMINI API ---
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "THAY_API_KEY_CỦA_BẠN_VÀO_ĐÂY_NẾU_KHÔNG_DÙNG_ENV")
genai.configure(api_key=GEMINI_API_KEY)

# --- ĐỊNH NGHĨA STRUCTURAL DATA MODELS (PYDANTIC) ---
class HookModel(BaseModel):
    formula_type: str
    hook_text: str
    video_visual: str
    static_layout: str
    copy_direction: str

class SmallIdeaModel(BaseModel):
    concept_code: str
    concept_name: str
    priority: str
    visual_direction: str
    suggested_hooks: List[HookModel]

class BigIdeaModel(BaseModel):
    idea_name: str
    small_ideas: List[SmallIdeaModel]

class PersonaModel(BaseModel):
    persona_name: str
    persona_code: str
    demographics: str
    description: str
    behavior: str
    insights: str
    motivation: str
    key_message: str
    suggested_formats: str
    big_ideas: List[BigIdeaModel]

class CreativeMatrixResponse(BaseModel):
    product_summary: str
    project_prefix: str
    personas: List[PersonaModel]

class NamingPayload(BaseModel):
    project_prefix: str
    audience_code: str
    audience_name: str
    concept_code: str
    concept_name: str
    idea_number: int
    format_code: str
    lang_code: str

# --- ENDPOINT TRANG CHỦ ---
@app.get("/")
def read_root():
    return {"status": "online", "message": "CreativeOps AI Engine đang hoạt động thành công!"}

# --- ENDPOINT 1: PHÂN RÃ MA TRẬN Ý TƯỞNG (GIẢI PHÁP CHUYÊN DỤNG CHO API MIỄN PHÍ) ---
@app.post("/api/v1/generate-matrix", response_model=CreativeMatrixResponse)
async def generate_matrix(product_description: str, target_market: str = "Vietnam"):
    system_instruction = """
    Bạn là một Giám đốc sáng tạo (Creative Director) và Growth Marketer lão luyện theo trường phái CreativeOps.
    Nhiệm vụ của bạn là phân tích bản mô tả sản phẩm thô và bẻ gãy nó thành một Ma trận Sáng tạo thực chiến (Creative Matrix).
    
    Hãy chia sản phẩm thành 3 nhóm Persona khách hàng (Audience) độc lập. Đối với MỖI nhóm Persona, bạn phải phân tích sâu sắc các khía cạnh cốt lõi sau để điền vào đúng cấu trúc JSON:
    - demographics: Độ tuổi, giới tính, phân khúc (Ví dụ: "18-32, All (~70% Female)").
    - description: Mô tả chi tiết hành vi, phong cách sống, gu thẩm mỹ (Ví dụ: "Thích nhập vai tình cảm, quen với webtoon...").
    - behavior: Thói quen tiêu thụ nội dung trên MXH (Ví dụ: "Thường xuyên đọc, xem hoặc chơi game liên quan...").
    - insights: Nỗi đau thầm kín hoặc khao khát sâu thẳm (Ví dụ: "Muốn cùng character viết ra 1 fantasy...").
    - motivation: Động lực cốt lõi thôi thúc họ hành động (Ví dụ: "Được sống trong một mối quan hệ...").
    - key_message: Thông điệp chí mạng (Ví dụ: "Câu chuyện của bạn. Nhân vật của bạn.").
    - suggested_formats: Các định dạng đề xuất (Ví dụ: "Static chat UI mockup, Video story reveal").

    Với mỗi Persona, hãy tạo ra 2 Big Ideas. Mỗi Big Idea có 2 Small Ideas (Concepts). Mỗi Small Idea sinh ra 2 Suggested Hooks cụ thể, thực chiến có kèm Layout ảnh tĩnh và Kịch bản video 3 giây đầu.
    Mã hóa project_prefix bằng 3 chữ cái viết hoa đại diện cho sản phẩm (Ví dụ sản phẩm Imely -> IML).

    BẮT BUỘC TRẢ VỀ THEO ĐÚNG ĐỊNH DẠNG JSON MẪU DƯỚI ĐÂY:
    {
      "product_summary": "Tóm tắt ngắn gọn thế mạnh sản phẩm tại đây",
      "project_prefix": "IML",
      "personas": [
        {
          "persona_name": "Tên nhóm khách hàng",
          "persona_code": "Mã (VD: YPI)",
          "demographics": "Thông tin nhân khẩu học",
          "description": "Mô tả hành vi",
          "behavior": "Thói quen MXH",
          "insights": "Insight sâu thẳm",
          "motivation": "Động lực mua hàng",
          "key_message": "Thông điệp chí mạng",
          "suggested_formats": "Định dạng khuyên dùng",
          "big_ideas": [
            {
              "idea_name": "Tên ý tưởng lớn",
              "small_ideas": [
                {
                  "concept_code": "Mã concept (VD: EA)",
                  "concept_name": "Tên concept cụ thể",
                  "priority": "High/Medium/Low",
                  "visual_direction": "Hướng hình ảnh",
                  "suggested_hooks": [
                    {
                      "formula_type": "Loại công thức hook",
                      "hook_text": "Tiêu đề gây chú ý 3s đầu",
                      "video_visual": "Kịch bản video",
                      "static_layout": "Bố cục ảnh tĩnh",
                      "copy_direction": "Hướng viết caption"
                    }
                  ]
                }
              ]
            }
          ]
        }
      ]
    }
    CHÚ Ý: Chỉ trả ra duy nhất khối JSON, không viết thêm lời mở đầu hay kết luận.
    """

    prompt = f"""
    Sản phẩm / Chiến dịch cần phân tích: {product_description}
    Thị trường mục tiêu: {target_market}
    """

    try:
        # Sử dụng mô hình 2.5 Flash tối ưu chi phí và tốc độ tốt nhất
        model = genai.GenerativeModel(model_name="gemini-2.5-flash")
        
        response = model.generate_content(
            [system_instruction, prompt],
            generation_config={
                "temperature": 0.4,
                "max_output_tokens": 8192
            }
        )
        
        # Kỹ thuật bóc tách JSON bằng Regex phòng vệ từ xa cho API miễn phí
        text_response = response.text.strip()
        match = re.search(r'\{.*\}', text_response, re.DOTALL)
        if match:
            clean_json_str = match.group(0)
        else:
            clean_json_str = text_response
            
        matrix_data = json.loads(clean_json_str)
        return matrix_data
    except Exception as e:
        print(f"❌ Lỗi Gemini rã ma trận: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Lỗi AI khởi tạo ma trận: {str(e)}")

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- ENDPOINT 2: XUẤT EXCEL CHUẨN DESIGN ĐẸP CAO CẤP ---
@app.post("/api/v1/export-matrix-excel")
async def export_matrix_excel(data: CreativeMatrixResponse):
    try:
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            attributes = ["DEMOGRAPHIC", "DESCRIPTION", "BEHAVIOR", "INSIGHTS", "MOTIVATION", "KEY MESSAGE", "FORMAT (gợi ý thôi)"]
            ta_dict = {"Thuộc Tính / Trường Dữ Liệu": attributes}
            
            for p in data.personas:
                col_name = f"{p.persona_name} ({p.persona_code})"
                ta_dict[col_name] = [
                    p.demographics, p.description, p.behavior, p.insights, p.motivation, p.key_message, p.suggested_formats
                ]
                
            df_ta = pd.DataFrame(ta_dict)
            df_ta.to_excel(writer, sheet_name="Target Audience", index=False)
            
            plan_rows = []
            global_hook_counter = 1
            
            for p in data.personas:
                for b_idea in p.big_ideas:
                    for s_idea in b_idea.small_ideas:
                        for h in s_idea.suggested_hooks:
                            idea_str_num = str(global_hook_counter).zfill(3)
                            creative_code = f"{data.project_prefix}-{p.persona_code}-{s_idea.concept_code}-{idea_str_num}-ST-VN"
                            
                            plan_rows.append({
                                "Creative Code": creative_code,
                                "Audience Persona": p.persona_name,
                                "Big Idea": b_idea.idea_name,
                                "Small Idea / Concept": s_idea.concept_name,
                                "Concept Code": s_idea.concept_code,
                                "Priority": s_idea.priority,
                                "Hook Formula": h.formula_type,
                                "Hook Text": h.hook_text,
                                "Visual Direction": s_idea.visual_direction,
                                "Video Visual (3s Đầu)": h.video_visual,
                                "Static Layout Composition": h.static_layout,
                                "Body Copy Direction": h.copy_direction
                            })
                            global_hook_counter += 1
                            
            df_plan = pd.DataFrame(plan_rows)
            df_plan.to_excel(writer, sheet_name="Creative Matrix Management", index=False)
            
            workbook = writer.book
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_body = Font(name="Segoe UI", size=10, bold=False, color="1E293B")
            fill_header = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid") 
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0')
            )
            
            ws1 = workbook["Target Audience"]
            ws1.row_dimensions[1].height = 28
            for col_idx in range(1, ws1.max_column + 1):
                cell = ws1.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
            
            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
                for cell in row:
                    cell.font = font_body
                    cell.border = thin_border
                    if cell.column == 1:
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="475569")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = align_left
            
            ws1.column_dimensions['A'].width = 25
            for col in range(2, ws1.max_column + 1):
                col_letter = get_column_letter(col)
                ws1.column_dimensions[col_letter].width = 40 
                
            ws2 = workbook["Creative Matrix Management"]
            ws2.row_dimensions[1].height = 28
            for col_idx in range(1, ws2.max_column + 1):
                cell = ws2.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
                for cell in row:
                    cell.font = font_body
                    cell.border = thin_border
                    cell.alignment = align_left
                    
            for col in ws2.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 45)

        output.seek(0)
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Creative_Ops_Matrix_{data.project_prefix}.xlsx"}
        )
    except Exception as e:
        print(f"❌ Lỗi xuất Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Lỗi hệ thống khi xuất file Excel: {str(e)}")

# --- ENDPOINT 3: BỘ DẬP MÃ NAMING TỰ ĐỘNG ---
@app.post("/api/v1/generate-dynamic-naming")
async def generate_dynamic_naming(payload: NamingPayload):
    idea_str = str(payload.idea_number).zfill(3)
    creative_code = f"{payload.project_prefix}-{payload.audience_code}-{payload.concept_code}-{idea_str}-{payload.format_code}-{payload.lang_code}"
    file_name_display = f"{payload.project_prefix} · {payload.audience_name} · {payload.concept_name} · Idea {idea_str} · Form: {payload.format_code} · [{payload.lang_code}]"
    return {"creative_code": creative_code.upper(), "file_name_display": file_name_display}

# --- ENDPOINT 4: AI PHÂN TÍCH HIỆU SUẤT DATA-DRIVEN ---
@app.post("/api/v1/analyze-performance")
async def analyze_performance(report_data_table: str):
    system_instruction = """
    Bạn là một Lead Data Analyst kiêm Performance Marketing Expert chuyên tối ưu tài nguyên Ads.
    Bạn sẽ nhận được bảng dữ liệu Markdown chứa: Mã Creative Code, Chi phí (Spend), Số click, Số đơn hàng (Conversions), CTR, CPA.
    Nhiệm vụ của bạn là bóc tách dữ liệu để phân loại: Nhóm tài nguyên nào là Winner (giữ lại, scale up), nhóm nào là Loser (tắt, đổi hook).
    
    Hãy phân tích và trả về cấu trúc JSON thuần có các keys sau: overall_evaluation, top_performing_persona, top_performing_concept, detailed_insights, next_action_steps.
    """
    try:
        model = genai.GenerativeModel(model_name="gemini-2.5-flash")
        response = model.generate_content([system_instruction, f"Dữ liệu báo cáo quảng cáo thực tế:\n{report_data_table}"])
        
        text_response = response.text.strip()
        match = re.search(r'\{.*\}', text_response, re.DOTALL)
        clean_json_str = match.group(0) if match else text_response
            
        return json.loads(clean_json_str)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi AI phân tích dữ liệu: {str(e)}")
